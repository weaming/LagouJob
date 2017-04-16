import logging
import os

from openpyxl import Workbook

logging.basicConfig(filename='info.log', level=logging.DEBUG)

EXCEL_DIR = 'D:/lagou/data/'


def write_excel(joblist, filename):
    mkdirs_if_not_exists(EXCEL_DIR)
    wb = Workbook()
    ws = wb.active
    ws.title = "职位信息"
    ws.cell(row=1, column=1).value = '职位编码'
    ws.cell(row=1, column=2).value = '职位名称'
    ws.cell(row=1, column=3).value = '所在城市'
    ws.cell(row=1, column=4).value = '发布日期'
    ws.cell(row=1, column=5).value = '薪资待遇'
    ws.cell(row=1, column=6).value = '公司编码'
    ws.cell(row=1, column=7).value = '公司名称'
    ws.cell(row=1, column=8).value = '公司全称'

    rownum = 2
    for each_job in joblist:
        ws.cell(row=rownum, column=1).value = each_job.positionId
        ws.cell(row=rownum, column=2).value = each_job.positionName
        ws.cell(row=rownum, column=3).value = each_job.city
        ws.cell(row=rownum, column=4).value = each_job.createTime
        ws.cell(row=rownum, column=5).value = each_job.salary
        ws.cell(row=rownum, column=6).value = each_job.companyId
        ws.cell(row=rownum, column=7).value = each_job.companyName
        ws.cell(row=rownum, column=8).value = each_job.companyFullName
        rownum += 1
    wb.save(EXCEL_DIR + filename + '.xlsx')
    logging.info('Excel生成成功!')


def get_mean(numstring):
    if '-' in numstring:
        min_str = numstring.split('-')[0]
        max_str = numstring.split('-')[1]
        return (float(min_str.strip().replace('k', '')) + float(max_str.strip().replace('k', ''))) / 2
    else:
        return float(numstring.replace('k', '').strip())


def mkdirs_if_not_exists(directory_):
    """create a new folder if it does not exist"""
    if not os.path.exists(directory_) or not os.path.isdir(directory_):
        os.makedirs(directory_)
